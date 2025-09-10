import os, sys
from pathlib import Path
import re
import asyncio
import tkinter as tk
from tkinter import filedialog
import pandas as pd
from openpyxl import load_workbook
from playwright.async_api import async_playwright

if getattr(sys, "frozen", False):
    os.environ["PLAYWRIGHT_BROWSERS_PATH"] = str(Path(sys._MEIPASS) / "ms-playwright")
else:
    os.environ.setdefault("PLAYWRIGHT_BROWSERS_PATH", "0")

# exe 파일 생성시 명령어

# $env:PLAYWRIGHT_BROWSERS_PATH = "0"
# >> & "C:\Users\Jaeyong.Choi\AppData\Local\Programs\Python\Python313\python.exe" -m playwright install chromium
# >> # => 현재 폴더에 .\ms-playwright 생성

# pyinstaller --onefile `
# >>   --add-data "$env:LOCALAPPDATA\ms-playwright;ms-playwright" `
# >>   ReceiptChecker.py

TARGET_PHRASE = "작성자가 삭제하거나 유효하지 않은 리뷰입니다."

# ---------- 1) 엑셀 파일 선택 ----------
root = tk.Tk(); root.withdraw()
excel_path = filedialog.askopenfilename(
    title="엑셀 파일을 선택하세요 (.xlsx 권장)",
    filetypes=[("Excel files", "*.xlsx")]
)
if not excel_path:
    raise SystemExit("파일을 선택하지 않았습니다.")

# ---------- 2) 엑셀 읽기 ----------
df = pd.read_excel(excel_path, sheet_name=0)
if df.shape[1] < 3:
    raise SystemExit("엑셀에 C열(세 번째 열)이 없습니다. URL이 C열에 있어야 합니다.")

urls_series = df.iloc[:, 2]  # C열 (URL)
total_rows = len(df)

# ---------- 3) openpyxl로 파일 열어 G열 쓰기 준비 ----------
wb = load_workbook(excel_path)
ws = wb.active  # 첫 번째 시트
if ws.cell(row=1, column=7).value in (None, ""):
    ws.cell(row=1, column=7).value = "상태"

# ---------- 4) Playwright로 렌더링 후 문구 확인 ----------
async def fetch_and_check(page, raw_url: str):
    if not isinstance(raw_url, str) or not raw_url.strip():
        return None
    url = raw_url.strip()
    if not re.match(r"^https?://", url, re.I):
        url = "http://" + url  # 스킴 보정

    resp = await page.goto(url, wait_until="domcontentloaded", timeout=30000)
    try:
        await page.wait_for_load_state("networkidle", timeout=15000)
    except Exception:
        pass

    # 지연 로딩 대비 스크롤
    for _ in range(5):
        await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
        await page.wait_for_timeout(700)

    html = await page.content()
    return TARGET_PHRASE in html

async def main():
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context(
            viewport={"width": 1366, "height": 768},
            locale="ko-KR",
            user_agent=("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                        "AppleWebKit/537.36 (KHTML, like Gecko) "
                        "Chrome/124.0.0.0 Safari/537.36")
        )
        page = await context.new_page()

        for i, raw_url in enumerate(urls_series, start=2):  # 2행부터 데이터 시작
            try:
                found = await fetch_and_check(page, str(raw_url) if pd.notna(raw_url) else "")
                if found:
                    ws.cell(row=i, column=7).value = "누락"
                    print(f"{i-1} [누락] {raw_url}")
                else:
                    print(f"{i-1} [정상] {raw_url}")
            except Exception as e:
                ws.cell(row=i, column=7).value = f"오류: {str(e)[:40]}"
                print(f"{i-1} [오류] {raw_url} -> {e}")

        await browser.close()

    wb.save(excel_path)
    print(f"\n엑셀 G열 업데이트 완료: {excel_path}")

if __name__ == "__main__":
    asyncio.run(main())